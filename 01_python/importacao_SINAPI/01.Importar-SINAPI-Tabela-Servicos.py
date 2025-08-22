import sys
import os, unicodedata
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import json
import io
import openpyxl
import re

# Força encoding UTF-8 nos prints e logs
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# Abas alvo para importação
ALVOS = {"ISD", "ICD", "CSD", "CCD", "ANALITICO"}
# UFs válidas
UF_SET = {
    "AC","AL","AM","AP","BA","CE","DF","ES","GO","MA","MG","MS","MT",
    "PA","PB","PE","PI","PR","RJ","RN","RO","RR","RS","SC","SE","SP","TO"
}

def log_importacao(msg, origem, usuario=None, ip=None):
    """
    Registra mensagem no log centralizado de importação SINAPI, incluindo data/hora, origem, usuário e IP.
    """
    # Tentar diferentes estratégias para encontrar o arquivo de log
    log_paths = []
    
    # Estratégia 1: Caminho relativo ao diretório atual
    log_paths.append(os.path.join('storage', 'logs', 'importacao_tabelas_oficiais.log'))
    
    # Estratégia 2: Caminho absoluto a partir do script Python
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(os.path.dirname(script_dir))
    log_paths.append(os.path.join(project_root, 'storage', 'logs', 'importacao_tabelas_oficiais.log'))
    
    # Estratégia 3: Caminho absoluto a partir do diretório de trabalho atual
    current_dir = os.getcwd()
    log_paths.append(os.path.join(current_dir, 'storage', 'logs', 'importacao_tabelas_oficiais.log'))
    
    # Estratégia 4: Caminho no diretório temporário do sistema
    import tempfile
    temp_dir = tempfile.gettempdir()
    log_paths.append(os.path.join(temp_dir, 'orcacidade_sinapi.log'))
    
    # Estratégia 5: Caminho local como último recurso
    log_paths.append('sinapi_importacao.log')
    
    log_path = None
    for path in log_paths:
        try:
            # Verificar se o diretório existe e criar se necessário
            log_dir = os.path.dirname(path)
            if log_dir and not os.path.exists(log_dir):
                try:
                    # Usar permissões mais permissivas para produção
                    os.makedirs(log_dir, mode=0o755, exist_ok=True)
                except (OSError, PermissionError):
                    continue  # Tentar próximo caminho se não conseguir criar diretório
            
            # Tentar abrir o arquivo para escrita (teste de permissão)
            try:
                with open(path, 'a', encoding='utf-8') as f:
                    pass
                log_path = path
                break
            except (OSError, PermissionError, IOError):
                continue
                
        except Exception:
            continue
    
    # Se encontrou um caminho válido, tentar escrever o log
    if log_path:
        try:
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            user_info = f" | Usuario: {usuario}" if usuario else ""
            ip_info = f" | IP: {ip}" if ip else ""
            
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(f"[{now}] [{origem}]{user_info}{ip_info} {msg}\n")
                f.flush()  # Força escrita imediata
                os.fsync(f.fileno())  # Sincroniza com disco (importante em produção)
                
        except Exception as e:
            # Se falhar o log, apenas ignorar (não quebrar o processamento)
            # Em produção, isso é crítico para não interromper o fluxo
            pass

def norm(s:str)->str:
    """
    Normaliza string para comparação robusta (remove acentos, espaços, caixa).
    """
    return unicodedata.normalize("NFKD",s).encode("ascii","ignore").decode().upper().replace(' ', '')

def is_empty(val) -> bool:
    """
    Verifica se valor é vazio ou nulo.
    """
    return pd.isna(val) or str(val).strip() == ''

def num_to_str(val, casas=2):
    """
    Converte número para string formatada com casas decimais, preservando zeros.
    """
    if is_empty(val):
        return ""
    try:
        f = float(str(val).replace(",", "."))
    except ValueError:
        return str(val).strip()
    s = f"{f:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return s

def codigo_to_str(val, largura=5):
    """
    Normaliza código para string com zeros à esquerda.
    """
    if is_empty(val):
        return ""
    try:
        i = int(float(str(val).replace(",", ".")))
    except ValueError:
        return str(val).strip()
    return str(i).zfill(largura)

def df_preco_only(arquivo, sheet, tipo, desoneracao):
    """
    Lê e processa as abas ISD, ICD, CSD, CCD do arquivo Excel, extraindo colunas fixas por letra e preenchendo campos padronizados.
    """
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)
    ws = wb[sheet]
    data_base = ws['B3'].value
    data_emissao = ws['B4'].value
    wb.close()

    # Ajusta data_base para formato '01/mm/yyyy' se vier como 'mm/yyyy'
    if data_base and re.match(r"^\d{2}/\d{4}$", str(data_base)):
        data_base = f"01/{data_base}"

    usecols = 'A,B,C,D,W'
    if tipo == 'insumos':
        colunas_saida = ['classificacao', 'codigo_insumo', 'descricao', 'unidade', 'custo_PR']
    elif tipo == 'composicoes':
        colunas_saida = ['grupo', 'codigo_composicao', 'descricao', 'unidade', 'custo_PR']
    else:
        raise ValueError(f"Tipo inválido: {tipo}")

    df = pd.read_excel(
        arquivo, sheet_name=sheet,
        header=9, usecols=usecols, dtype=str, engine="openpyxl"
    )
    df.columns = colunas_saida
    df['custo_PR'] = df['custo_PR'].apply(num_to_str, casas=2)
    df['data_base'] = data_base
    df['data_emissao'] = data_emissao
    df['desoneracao'] = desoneracao
    return df[['classificacao' if tipo=='insumos' else 'grupo',
               'codigo_insumo' if tipo=='insumos' else 'codigo_composicao',
               'descricao', 'unidade', 'custo_PR', 'data_base', 'data_emissao', 'desoneracao']]

def processa_insumos(arquivo, sheet, desoneracao, usuario=None, ip=None):
    """
    Processa aba de insumos (ISD/ICD), retornando DataFrame padronizado e logando início/fim.
    """
    origem = f"SINAPI_{sheet.upper()}"
    log_importacao(f"Importação iniciada para aba {sheet} do arquivo {os.path.basename(arquivo)}", origem, usuario, ip)
    df = df_preco_only(arquivo, sheet, tipo='insumos', desoneracao=desoneracao)
    log_importacao(f"{len(df)} insumos importados com sucesso.", origem, usuario, ip)
    return df

def processa_compos(arquivo, sheet, desoneracao, usuario=None, ip=None):
    """
    Processa aba de composições (CSD/CCD), retornando DataFrame padronizado e logando início/fim.
    """
    origem = f"SINAPI_{sheet.upper()}"
    log_importacao(f"Importação iniciada para aba {sheet} do arquivo {os.path.basename(arquivo)}", origem, usuario, ip)
    df = df_preco_only(arquivo, sheet, tipo='composicoes', desoneracao=desoneracao)
    log_importacao(f"{len(df)} composições importadas com sucesso.", origem, usuario, ip)
    return df

def processa_analitico(arquivo, sheet, usuario=None, ip=None):
    """
    Processa a aba Analítico, identifica composições pai e itens, registra logs detalhados e retorna DataFrame.
    """
    origem = "SINAPI_ANALITICO"
    try:
        log_importacao(f"Importação iniciada para aba Analítico do arquivo {os.path.basename(arquivo)}", origem, usuario, ip)
        df = pd.read_excel(
            arquivo, sheet_name=sheet,
            header=9, usecols="A:H", dtype=str, engine="openpyxl",
            names=["grupo", "codigo_composicao", "tipo_item", "codigo_item",
                   "descricao", "unidade", "coeficiente", "situacao"]
        )
        df["codigo_composicao"] = df["codigo_composicao"].apply(codigo_to_str)
        df["codigo_item"] = df["codigo_item"].apply(codigo_to_str)
        df["coeficiente"] = df["coeficiente"].apply(lambda x: num_to_str(x, 5))
        # Identifica composições pai e itens (considera vazio ou NaN)
        is_pai = (df["tipo_item"].isna() | (df["tipo_item"] == "")) & (df["codigo_item"].isna() | (df["codigo_item"] == ""))
        n_pai = is_pai.sum()
        n_itens = (~is_pai).sum()
        log_importacao(f"{n_pai} composições pai e {n_itens} itens importados com sucesso.", origem, usuario, ip)
        return df
    except Exception as e:
        log_importacao(f"Erro na importação da aba Analítico: {str(e)}", origem, usuario, ip)
        raise

def processar_arquivo(caminho_arquivo, diretorio_saida, usuario=None, ip=None):
    """
    Função principal: processa todas as abas do arquivo Excel, exporta arquivos processados e metadados.
    A limpeza dos arquivos temporários deve ser feita pelo backend após o uso pelo frontend.
    """
    try:
        wb = load_workbook(caminho_arquivo, read_only=True, data_only=True)
        abas_exist = {norm(n): n for n in wb.sheetnames}
        abas_proc = [abas_exist[x] for x in ALVOS if x in abas_exist]
        if not abas_proc:
            raise ValueError("Nenhuma aba alvo encontrada!")

        mes_referencia = wb[abas_proc[0]]["B3"].value
        data_emissao = wb[abas_proc[0]]["B4"].value
        wb.close()

        os.makedirs(diretorio_saida, exist_ok=True)

        for aba in abas_proc:
            n = norm(aba)
            if n == "ISD":
                df = processa_insumos(caminho_arquivo, aba, desoneracao="sem", usuario=usuario, ip=ip)
            elif n == "ICD":
                df = processa_insumos(caminho_arquivo, aba, desoneracao="com", usuario=usuario, ip=ip)
            elif n == "CSD":
                df = processa_compos(caminho_arquivo, aba, desoneracao="sem", usuario=usuario, ip=ip)
            elif n == "CCD":
                df = processa_compos(caminho_arquivo, aba, desoneracao="com", usuario=usuario, ip=ip)
            elif n == "ANALITICO":
                df = processa_analitico(caminho_arquivo, aba, usuario=usuario, ip=ip)
            else:
                continue
            # Exporta apenas as colunas padrão (já garantido nas funções)
            arquivo_saida = os.path.join(diretorio_saida, f'sinapi_processado_{aba}.xlsx')
            df.to_excel(arquivo_saida, index=False)

        with open(os.path.join(diretorio_saida, 'sinapi_metadata.json'), 'w') as f:
            json.dump({
                'mes_referencia': mes_referencia,
                'data_emissao': str(data_emissao),
                'abas_processadas': abas_proc
            }, f)

        return 0

    except Exception as e:
        print(str(e), file=sys.stderr)
        return 1
    # A limpeza dos arquivos temporários deve ser feita pelo backend Laravel após o uso pelo frontend.

if __name__ == "__main__":
    # Permite receber usuário e IP como argumentos opcionais
    if len(sys.argv) < 3:
        print("Uso: script.py <entrada.xlsx> <diretorio_saida> [usuario] [ip]", file=sys.stderr)
        sys.exit(1)

    caminho_arquivo = sys.argv[1]
    diretorio_saida = sys.argv[2]
    usuario = sys.argv[3] if len(sys.argv) > 3 else None
    ip = sys.argv[4] if len(sys.argv) > 4 else None
    sys.exit(processar_arquivo(caminho_arquivo, diretorio_saida, usuario, ip))
